"""This script reads in a ledger and actual account balance to generate a report"""
# Beta version
#################################################################################################
# 2/1/2019 Constructed main program to automatically read in account names and basic start info
# 6/3/2019 Changed method of reading in historical data. Read in previous year and make changes
#          to parent directory file.
# 6/17/2019 Simplified wording of transfer PMTs into and out of CapCom accounts. Took out FN.
#################################################################################################

import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import time
from datetime import date
today = date.today()

# INPUT CONSTANTS
LONG_TERM_PRICE_INC = 2.8              # long term pricing in % (Source: Florence S.). Default 2.8.
TRANSFER_IN_PMT = 0             # Enter in total amount transferred into CapCom accounts
TRANSFER_OUT_PMT = 0          # Enter in total amount transferred out of CapCom accounts
ECS = 0                                # East Commerce Solutions overPMT. waiting for reimbursement.

HSY = 1997                              # Historical results start year. Should be 1997
SPY = 2009                              # Starting plot year. Default 2009.

# You need to generate a separate folder by year. Note, this is done autom in the startfile.
DIRECTORY_PATH = os.path.dirname(os.path.realpath(__file__))

# First read master start file continuing correct year to create new folder and files as
# well as a current account list.
STARTFILE = pd.ExcelFile(DIRECTORY_PATH + '/MasterStartFile.xlsx')
A = STARTFILE.sheet_names
DF1 = STARTFILE.parse(A[0])

YEAR = int(DF1['Year'][0])
DATES = (str(DF1['Days'][0][1:-1]), str(DF1['Days'][1][1:-1]), str(DF1['Days'][2][1:-1]))
ACCOUNTS = DF1['Accounts']
ACCOUNTS_NUM = len(DF1['Accounts'])

os.chdir(DIRECTORY_PATH)

os.chdir("..")
DIRECTORY_MAIN_PATH = os.getcwd()
os.chdir(DIRECTORY_PATH)

plt.close("all")

pd.options.display.float_format = '${:,.2f}'.format

# Load the main festival ledger. This contains all transactions from the festival.
# Note that the individual collections from festival days are summed from main ledger
# used during the actual festival days.

XL = pd.ExcelFile(DIRECTORY_PATH+'/ledger_' + str(YEAR) +'.xlsx')
# Print the sheet names
print(XL.sheet_names)
A = XL.sheet_names
# Load a sheet into a DataFrame by name: df1
DF1 = XL.parse(A[0])

FL = pd.read_csv(DIRECTORY_PATH+'/ledgerFestival_'+str(YEAR) +'.csv')

# Build up a summary table from festival days transactions. Includes credit card charge corrections
FEST_DAYS_SUMMARY = []

for i in range(0, ACCOUNTS_NUM):
    INCOME_FEST_DAYS = FL[(FL['account'] == ACCOUNTS[i]) & (FL['type'] == "income")]['amount'].sum() -  FL[(FL['account'] == ACCOUNTS[i]) & (FL['type'] == "bank")]['amount'].sum()
    EXPENSE_FEST_DAYS = FL[(FL['account'] == ACCOUNTS[i]) & (FL['type'] == "expense")]['amount'].sum()
    FEST_DAYS_SUMMARY.append((ACCOUNTS[i], INCOME_FEST_DAYS, EXPENSE_FEST_DAYS, INCOME_FEST_DAYS
                              - EXPENSE_FEST_DAYS))

FEST_DAYS_SUMMARY = pd.DataFrame(FEST_DAYS_SUMMARY, columns=('Account', 'INCOME', 'EXPENSE', 'Profit'))

# write in values for historical file

HFF = ("souvlaki", "beverage out", "loukoumathes", "beverage in", "salads", "meals", "pastry", "arts and crafts", "vendor")
EC = ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O",
      "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AD", "AE",
      "AF", "AO", "AP", "AQ")

os.chdir(DIRECTORY_MAIN_PATH + '/' +str(YEAR-1))
H = load_workbook(filename='festival_historical.xlsx')
H1 = H.worksheets[0]

for i in range(0, 9):

    H1[EC[(3*i)+4] + str(YEAR - HSY)].value = FL[(FL['account'] == HFF[i]) & (FL['type'] == "income") & (FL['date'] == DATES[0])]['amount'].sum() -  FL[(FL['account'] == HFF[i]) & (FL['type'] == "bank") & (FL['date'] == DATES[0])]['amount'].sum() 
    H1[EC[(3*i)+5] + str(YEAR - HSY)].value = FL[(FL['account'] == HFF[i]) & (FL['type'] == "income") & (FL['date'] == DATES[1])]['amount'].sum() -  FL[(FL['account'] == HFF[i]) & (FL['type'] == "bank") & (FL['date'] == DATES[1])]['amount'].sum()
    H1[EC[(3*i)+6] + str(YEAR - HSY)].value = FL[(FL['account'] == HFF[i]) & (FL['type'] == "income") & (FL['date'] == DATES[2])]['amount'].sum() -  FL[(FL['account'] == HFF[i]) & (FL['type'] == "bank") & (FL['date'] == DATES[2])]['amount'].sum()

H1[EC[0] + str(YEAR - HSY)].value =  H1['F' + str(YEAR - HSY)].value + H1['I' + str(YEAR - HSY)].value + H1['L' + str(YEAR - HSY)].value + H1['O' + str(YEAR - HSY)].value + H1['R' + str(YEAR - HSY)].value + H1['U' + str(YEAR - HSY)].value + H1['X' + str(YEAR - HSY)].value
H1[EC[1] + str(YEAR - HSY)].value =  H1['G' + str(YEAR - HSY)].value + H1['J' + str(YEAR - HSY)].value + H1['M' + str(YEAR - HSY)].value + H1['P' + str(YEAR - HSY)].value + H1['S' + str(YEAR - HSY)].value + H1['V' + str(YEAR - HSY)].value + H1['Y' + str(YEAR - HSY)].value
H1[EC[2] + str(YEAR - HSY)].value =  H1['H' + str(YEAR - HSY)].value + H1['K' + str(YEAR - HSY)].value + H1['N' + str(YEAR - HSY)].value + H1['Q' + str(YEAR - HSY)].value + H1['T' + str(YEAR - HSY)].value + H1['W' + str(YEAR - HSY)].value + H1['Z' + str(YEAR - HSY)].value

H1[EC[3] + str(YEAR - HSY)].value = H1['B' + str(YEAR - HSY)].value + H1['C' + str(YEAR - HSY)].value + H1['D' + str(YEAR - HSY)].value
H1[EC[3] + str(YEAR - HSY)].value = H1['B' + str(YEAR - HSY)].value + H1['C' + str(YEAR - HSY)].value + H1['D' + str(YEAR - HSY)].value
H1['AG' + str(YEAR - HSY)].value = H1['AD' + str(YEAR - HSY)].value + H1['AE' + str(YEAR - HSY)].value + H1['AF' + str(YEAR - HSY)].value

H1["A" + str(YEAR-HSY)].value = YEAR

H.save(DIRECTORY_PATH + '/festival_historical.xlsx')

ANALYSIS_FESTIVAL_DATE = DATES[1]

ML = load_workbook(filename=DIRECTORY_PATH+'/ledger_'+str(YEAR)+'.xlsx')
ML1 = ML.worksheets[0]

for i in range(0, ACCOUNTS_NUM):
    ML1["A" + str(i + 2)].value = ""
    ML1["A" + str(i + ACCOUNTS_NUM + 2)].value = ""
    ML1["B" + str(i + 2)].value = ANALYSIS_FESTIVAL_DATE
    ML1["B" + str(i + ACCOUNTS_NUM + 2)].value = ANALYSIS_FESTIVAL_DATE
    ML1["C" + str(i + 2)].value = "income"
    ML1["C" + str(i + ACCOUNTS_NUM + 2)].value = "expense"
    ML1["D" + str(i + 2)].value = FEST_DAYS_SUMMARY['Account'][i]
    ML1["D" + str(i + ACCOUNTS_NUM + 2)].value = FEST_DAYS_SUMMARY['Account'][i]
    ML1["E" + str(i + 2)].value = FEST_DAYS_SUMMARY['INCOME'][i]
    ML1["E" + str(i + ACCOUNTS_NUM + 2)].value = FEST_DAYS_SUMMARY['EXPENSE'][i]
    ML1["F" + str(i + 2)].value = "y"
    ML1["F" + str(i + ACCOUNTS_NUM + 2)].value = "y"
    ML1["M" + str(i + 2)].value = "y"
    ML1["M" + str(i + ACCOUNTS_NUM + 2)].value = "y"

ML.save(DIRECTORY_PATH + '/ledger_' + str(YEAR) + '.xlsx')

XL = pd.ExcelFile(DIRECTORY_PATH + '/ledger_' + str(YEAR) + '.xlsx')

# Print the sheet names
print(XL.sheet_names)
A = XL.sheet_names
# Load a sheet into a DataFrame by name: df1
DF1 = XL.parse(A[0])

EXPENSES = DF1[DF1['Type'] == "expense"]['Amount'].sum()
PAYABLES = DF1[(DF1['Type'] == "expense") & (DF1['Reconciled'] != "y")]['Amount'].sum()
INCOME = DF1[DF1['Type'] == "income"]['Amount'].sum()
RECEIVABLES = DF1[(DF1['Type'] == "income") & (DF1['Reconciled'] != "y")]['Amount'].sum()

ACTUAL_RECONCILE = INCOME - EXPENSES + PAYABLES - RECEIVABLES
CAPEX = DF1[DF1['Account'] == "capital expense"]['Amount'].sum()

CAPEX_AND_INVENTORY = DF1[(DF1['Account'] == "capital expense") | (DF1['Account'] == "inventory")]['Amount'].sum()

# Read in the CapCom transaction summary. Use online account tool.
CC = pd.read_csv(DIRECTORY_PATH+'/CapComLedger.csv', skiprows=2)
AL = list(CC)
DATE_RANGE = AL[0]
CC = pd.read_csv(DIRECTORY_PATH + '/CapComLedger.csv', skiprows=3)
CURRENT_BALANCE = CC['Balance'].iloc[0]
START_BALANCE_DATE = CC['Date'].iloc[-1]
CURRENT_BALANCE_DATE = CC['Date'].iloc[0]
START_BALANCE = CC['Balance'].iloc[-1]

# Read in the CapCom Savings account transaction summary. Use online account tool.
CC_SAV = pd.read_csv(DIRECTORY_PATH + '/CapComLedgerSavings.csv', skiprows=2)
AL_SAV = list(CC_SAV)
DATE_RANGE_SAV = AL_SAV[0]
CC_SAV = pd.read_csv(DIRECTORY_PATH + '/CapComLedgerSavings.csv', skiprows=3)
CURRENT_BALANCE_SAV = CC_SAV['Balance'].iloc[0]
START_BALANCE_DATE_SAV = CC_SAV['Date'].iloc[-1]
CURRENT_BALANCE_DATE_SAV = CC_SAV['Date'].iloc[0]
START_BALANCE_SAV = CC_SAV['Balance'].iloc[-1]

# Quantify the sum of dividend payments in CapCom accounts
DIVIDEND = CC[CC['Description'] == "Deposit Dividend DIVIDEND"]['Amount Credit'].sum()
DIVIDEND_SAV = CC_SAV[CC_SAV['Description'] == "Deposit Dividend DIVIDEND"]['Amount Credit'].sum()

# Group by account. This will be used to generate main summary table.
A1 = DF1.groupby(['Type', 'Account'])['Amount'].sum()
B1 = A1.reset_index()
D1 = B1[B1['Amount'] != 0]
D1E = D1[D1['Type'] == 'expense']
D1I = D1[D1['Type'] == 'income']

D1I = D1I.drop(['Type'], axis=1)
D1E = D1E.drop(['Type'], axis=1)

C1 = DF1[DF1['FestivalDays'] == "y"]

FESTIVAL_FOOD_INC = B1.loc[(B1['Type'] == "income")
                           & ((B1['Account'] == "meals") | (B1['Account'] == "beverage in") | (B1['Account'] == "beverage out")
                              | (B1['Account'] == "souvlaki") | (B1['Account'] == "loukoumathes")
                              | (B1['Account'] == "pastry"))]['Amount'].sum()

# Main report is printed out. This includes a breakdown of all accounts, a main financial summary
# and a reconcliation of the accounts with the festival finance numbers.

print("---------------------------------------------------------------------")
print(YEAR, " Festival Finance Summary")
print("Date of analysis : ",today)
print(DATE_RANGE, "(Accounts reconcile analysis date range)")
print("---------------------------------------------------------------------")
print("Summary of Income")
print(D1I)
print("---------------------------------------------------------------------")
print("Summary of Expenses")
print(D1E)
print("---------------------------------------------------------------------")
print("Main Finance Summary")
print("---------------------------------------------------------------------")
print("Gross =                                      ", '${:,.2f}'.format(INCOME))
print("Expenses (Not including Capex) =             ", '${:,.2f}'.format(EXPENSES - CAPEX))
print("Expenses (Not including Capex or inventory) =", '${:,.2f}'.format(EXPENSES - CAPEX_AND_INVENTORY))
print("Expenses (Capex) =                           ", '${:,.2f}'.format(CAPEX))
print("Profit (Not including Capex) =               ", '${:,.2f}'.format(INCOME - EXPENSES + CAPEX))
print("Profit (Not including Capex or inventory) =  ", '${:,.2f}'.format(INCOME - EXPENSES + CAPEX_AND_INVENTORY))
print("Profit (Including Capex) =                   ", '${:,.2f}'.format(INCOME - EXPENSES))
print("Food & beverage sales during festival =      ", '${:,.2f}'.format(FESTIVAL_FOOD_INC))
print("---------------------------------------------------------------------")
print("Accounts Reconcile")
print("---------------------------------------------------------------------")
print("Analysis from festival transaction records")
print("    Total Income Reported =                  ", '${:,.2f}'.format(INCOME))
print("        MINUS Expenses =                     ", '${:,.2f}'.format(-EXPENSES))
print("        PLUS Payables  =                     ", '${:,.2f}'.format(PAYABLES))
print("        MINUS Receivables =                  ", '${:,.2f}'.format(-RECEIVABLES))
print("---------------------------------------------------------------------")
print("Expected increase/decrease in accounts =     ", '${:,.2f}'.format(INCOME - EXPENSES
                                                                         + PAYABLES - RECEIVABLES))
print("---------------------------------------------------------------------")
print("Checking Account info:")
print("    CapCom Organizational checking account:       XXXXX401-90")
print("    CapCom Organizational savings account:        XXXXX401-13")
print("---------------------------------------------------------------------")
print("CapCom transaction", DATE_RANGE)
print("---------------------------------------------------------------------")
print("Current CapCom Checking balance =            ", '${:,.2f}'.format(CURRENT_BALANCE))
print("Current CapCom Savings balance =             ", '${:,.2f}'.format(CURRENT_BALANCE_SAV))
print("    CURRENT CUMULATIVE ACCOUNT BALANCE =     ", '${:,.2f}'.format(CURRENT_BALANCE
                                                                         + CURRENT_BALANCE_SAV))
print("---------------------------------------------------------------------")
print("Inital CapCom Checking balance =             ", '${:,.2f}'.format(START_BALANCE))
print("Inital CapCom Savings balance =              ", '${:,.2f}'.format(START_BALANCE_SAV))
print("    BEGINNING CUMULATIVE ACCOUNT BALANCE  =  ", '${:,.2f}'.format(START_BALANCE
                                                                         + START_BALANCE_SAV))
print("---------------------------------------------------------------------")
print("Difference Current total from Initial total  ", '${:,.2f}'.format(CURRENT_BALANCE
                                                                         + CURRENT_BALANCE_SAV
                                                                         - START_BALANCE
                                                                         - START_BALANCE_SAV))
print("     PLUS xfer CapCom Checking --> Savings = ", '${:,.2f}'.format(0))
print("     MINUS TRANSFER INTO CapCom Account  =   ", '${:,.2f}'.format(TRANSFER_IN_PMT))
print("     PLUS TRANSFER OUT OF CapCom Account  =  ", '${:,.2f}'.format(TRANSFER_OUT_PMT))
print("     PLUS East Commerce Soln Overpay (IP) =  ", '${:,.2f}'.format(ECS))
print("     MINUS CapCom dividend PMTs =            ", '${:,.2f}'.format(-DIVIDEND - DIVIDEND_SAV))
print("---------------------------------------------------------------------")
print("Increase in account balances =               ", '${:,.2f}'.format(CURRENT_BALANCE
                                                                         + CURRENT_BALANCE_SAV
                                                                         - START_BALANCE
                                                                         - START_BALANCE_SAV
                                                                         - TRANSFER_IN_PMT
                                                                         + TRANSFER_OUT_PMT
                                                                         + ECS
                                                                         - DIVIDEND
                                                                         - DIVIDEND_SAV + 5))
print("---------------------------------------------------------------------")
print("Accounts Reconcile (should be zero) =        ", '${:,.2f}'.format(INCOME - EXPENSES
                                                                         + PAYABLES - RECEIVABLES - CURRENT_BALANCE
                                                                         - CURRENT_BALANCE_SAV
                                                                         + START_BALANCE
                                                                         + START_BALANCE_SAV + TRANSFER_IN_PMT
                                                                         - TRANSFER_OUT_PMT
                                                                         - ECS
                                                                         + DIVIDEND
                                                                         + DIVIDEND_SAV - 5))


H = load_workbook(filename=DIRECTORY_PATH +'/festival_historical.xlsx')
H1 = H.worksheets[0]

H1['AH' + str(YEAR - HSY)].value = B1.loc[(B1['Account'] == "$1 raffle") & (B1['Type'] == "income")]['Amount'].sum()
H1['AK' + str(YEAR - HSY)].value = B1.loc[(B1['Account'] == "$100 raffle") & (B1['Type'] == "income")]['Amount'].sum()

H1['AY' + str(YEAR - HSY)].value = INCOME
H1['AZ' + str(YEAR - HSY)].value = EXPENSES - CAPEX_AND_INVENTORY
H1['BA' + str(YEAR - HSY)].value = INCOME - (EXPENSES - CAPEX_AND_INVENTORY)

H1['BB' + str(YEAR - HSY)].value = DATES[0]
H1['BC' + str(YEAR - HSY)].value = DATES[1]
H1['BD' + str(YEAR - HSY)].value = DATES[2]

H.save(DIRECTORY_PATH+'/festival_historical.xlsx')

FS = pd.ExcelFile(DIRECTORY_PATH+'/festival_historical.xlsx')
FST = FS.sheet_names
# Load a sheet into a DataFrame by renaming FS
FS = FS.parse(FST[0])

FR = load_workbook(DIRECTORY_PATH+'/Final_'+ str(YEAR)+'_report_supporting_data.xlsx')
FR1 = FR.worksheets[0]

FR1['C4'].value = FS['total.fri'].iloc[-1]
FR1['C5'].value = FS['total.sat'].iloc[-1]
FR1['C6'].value = FS['total.sun'].iloc[-1]
FR1['C7'].value = FS['total.fri'].iloc[-1] + FS['total.sat'].iloc[-1] + FS['total.sun'].iloc[-1]
FR1['C8'].value = INCOME
FR1['C9'].value = EXPENSES - CAPEX_AND_INVENTORY
FR1['C10'].value = INCOME - EXPENSES + CAPEX_AND_INVENTORY
FR1['B12'].value = "Simulation run"
FR1['C12'].value = time.strftime("%c")
FR.save(DIRECTORY_PATH+'/Final_'+ str(YEAR)+'_report_supporting_data.xlsx')


plt.figure()

plt.plot(FS['year'], FS['total'], 'k', label='total')
plt.ylabel('$')
plt.xlabel('year')
plt.plot(FS['year'], FS['total.fri'], 'b--', label='fri')
plt.plot(FS['year'], FS['total.sat'], 'r--', label='sat')
plt.plot(FS['year'], FS['total.sun'], 'g--', label='sun')
plt.legend(loc='best')
plt.xlim(SPY, YEAR)
plt.savefig(DIRECTORY_PATH+'/FestDays.png')

plt.figure()

plt.plot(FS['year'], FS['gross'], 'k', label='sales')
plt.ylabel('$')
plt.xlabel('year')
plt.plot(FS['year'], FS['expenses'], 'r', label='expenses')
plt.plot(FS['year'], FS['profit'], 'g', label='profit')
plt.legend(loc='best')
plt.xlim(SPY, YEAR)
plt.savefig(DIRECTORY_PATH+'/MainMetrics.png')

plt.figure()

plt.plot(FS['year'], FS['gross'], linestyle=':', label='sales', color='black')
plt.ylabel('$')
plt.xlabel('year')
plt.plot(FS['year'], FS['expenses'], linestyle='--', label='expenses', color='black')
plt.plot(FS['year'], FS['profit'], 'k', label='profit', color='black')
plt.legend(loc='best')
plt.xlim(SPY, YEAR)
plt.savefig(DIRECTORY_PATH + '/MainMetricsBW.png')

plt.figure()

plt.plot(FS['year'], FS['total'], 'k')
plt.ylabel('3-Day Total Food Sales $')
plt.xlabel('year')
plt.xlim(SPY, YEAR)
plt.title('Food Sales')
plt.savefig(DIRECTORY_PATH+'/FoodSales.png')

plt.figure()
B = 100*FS['total'].pct_change()
#B1 = B[-8:]
B1 = B[-(YEAR-SPY):]
B2 = B1.mean()

plt.plot(FS['year'], B, 'k')
plt.ylabel('YoY %')
plt.xlabel('year')
plt.xlim(SPY + 1, YEAR)
plt.ylim(-20, 20)
plt.title('YoY increase in food sales')

plt.hlines(y=B2, xmin=SPY+1, xmax=YEAR, color='r', linestyle='--')
plt.hlines(y=LONG_TERM_PRICE_INC, xmin=SPY+1, xmax=YEAR, color='g', linestyle='--')
plt.text(YEAR - 4, 3.2, 'Long term festival average', color='r', fontsize=10)
plt.text(YEAR - 4, 0, 'Long term food price increases', color='g', fontsize=10)
plt.savefig(DIRECTORY_PATH+'/YoY.png')

plt.figure()

plt.plot(FS['year'], B, 'k')
plt.ylabel('YoY %')
plt.xlabel('year')
plt.xlim(SPY + 1, YEAR)
plt.ylim(-20, 20)
plt.title('YoY increase in food sales')

plt.hlines(y=B2, xmin=SPY+1, xmax=YEAR, color='black', linestyle='--')
plt.hlines(y=LONG_TERM_PRICE_INC, xmin=SPY+1, xmax=YEAR, color='black', linestyle=':')
plt.text(YEAR - 4, 3.2, 'Long term festival average', color='black', fontsize=10)
plt.text(YEAR - 4, 0, 'Long term food price increases', color='black', fontsize=10)
plt.savefig(DIRECTORY_PATH+'/YoYBW.png')

print("---------------------------------------------------------------------")
print("SYSTEM CHECKS                                                     ")
print("---------------------------------------------------------------------")
print("Are all ledger accounts contained within the main account list?")
print(set(DF1['Account']).issubset(ACCOUNTS))
print("Are all festival ledger accounts contained in the main ledger?")
print(set(FL['account']).issubset(DF1['Account']))


print("---------------------------------------------------------------------")
print("FESTIVAL RECORDS                                                     ")
print("---------------------------------------------------------------------")
print("Record FOOD SALES:    ", '${:,.2f}'.format(FS.loc[FS['total'].idxmax()]['total']), "(", FS.loc[FS['total'].idxmax()]['year'], ")")
print("Record GROSS:         ", '${:,.2f}'.format(FS.loc[FS['gross'].idxmax()]['gross']), "(", FS.loc[FS['gross'].idxmax()]['year'], ")")
print("Record EXPENSES:      ", '${:,.2f}'.format(FS.loc[FS['expenses'].idxmax()]['expenses']), "(", FS.loc[FS['expenses'].idxmax()]['year'], ")")
print("Record PROFIT:        ", '${:,.2f}'.format(FS.loc[FS['profit'].idxmax()]['profit']), "(", FS.loc[FS['profit'].idxmax()]['year'], ")")
print("---------------------------------------------------------------------")
